#! -*- coding:utf-8 -*-

################################################################################
#   程序功能：
#       多轮对话的数据辅助函数
################################################################################

from openpyxl import load_workbook
from openpyxl import Workbook
import os

#蠢回答列表
dull_answers = ["我不知道你在说什么",
                "我不清楚",
                "我不明白",
                "我也没有答案",
                "我不知道"]

################################################################################
# 获取数据函数，数据格式为：
#    会话集(列表):会话1(列表):语句1(字符串)
#                          语句2(字符串)
#                          语句3(字符串)
#                          语句4(字符串)
#                          问句5(列表):候选问句1(字典):
#                                        问句(键,字串)=>候选答案(值,列表):
#                                                                答案1(字串)
#                                                                   .
#                                                                   .
#                                                                   .
#                                                                答案m(字串)
#                                     候选问句2(字典)
#                                     候选问句3(字典)
#                                     候选问句4(字典)
#                                     候选问句5(字典)
#                           回答6(字符串)
#                 会话2
#                   .
#                   .
#                   .
#                 会话n
################################################################################
turn_per_epi = 6
candi_quests_num = 5
def fetch_episodes(file_name):
	# 提取后存储的数据文件
	data_file_name = file_name.replace('.xlsx', '.hd5')
	if os.path.exists(data_file_name):
		pass
	else:
		epi_line_num = 21961
		candi_line_num = 31815
		workbook = load_workbook(file_name)
		episode_sheet = workbook['episode']
		candidates_sheet = workbook['candidates']
		episodes = []
		epi_line = 2
		candi_line = 2
		# 遍历episode表中每一行，找到C列为5的行
		while epi_line <= epi_line_num:
			loc_epi = 'C' + str(epi_line)
			# 定位到每个episode的最后一句
			if int(episode_sheet[loc_epi].value) == 5:
				loc_epi = 'D' + str(epi_line-1)
				question = episode_sheet[loc_epi].value
				# 判断此问题是否有候选问句
				loc_candi = 'D' + str(candi_line)
				if question == candidates_sheet[loc_candi].value:
					episode = []
					# 将前四句对话放入片段中
					for i in range(epi_line-5, epi_line-1):
						loc_epi = 'D' + str(i)
						episode.append(episode_sheet[loc_epi].value)
					# 第五句为当前的问句，寻找其候选句子和候选答案
					questions = []
					questions.append(question)
					candidates = {}
					candi_quest = ""
					# 对所有候选问句进行循环
					while (question == candidates_sheet[loc_candi].value) and (candi_line <= candi_line_num):
						loc_candi_quest = 'F' + str(candi_line)
						if candi_quest != candidates_sheet[loc_candi_quest].value:
							candi_quest = candidates_sheet[loc_candi_quest].value
							candidates[candi_quest] = []
						loc_candi_ans = 'J' +str(candi_line)
						candi_ans = candidates_sheet[loc_candi_ans].value
						candidates[candi_quest].append(candi_ans)
						candi_line += 1
						loc_candi = 'D' + str(candi_line)
					questions.append(candidates)
					episode.append(questions)
					# 将答案放入episode中
					loc_epi = 'D' + str(epi_line)
					episode.append(episode_sheet[loc_epi].value)
					episodes.append(episode)
			epi_line += 1
	return episodes


# 展示episode的函数
state_num = 4
def display_data(episode):
    # 展示上下文
    print "State:",
    for i in range(state_num):
        print episode[i],
    print

    # 展示问句
    print "Query:", episode[4][0]

    # 展示候选问句
    for candi_quest in episode[4][1].keys():
        print "Candidate Question:", candi_quest,
        # 展示候选答案
        print "candidate answer:",
        for candi_ans in episode[4][1][candi_quest]:
            print candi_ans,
        print


# 上面的函数是在原始的multi_turn_data文件中使用
# 下面的函数则在后来的multi_turn_chat_data文件中使用
# 从文件中获取对话
def fetch_chats(file_name, sheet_name):
	lines_num = 235142
	workbook = load_workbook(file_name)
	worksheet = workbook[sheet_name]
	row = 2
	chats = []
	while row < lines_num:
		chat = []
		chatid = worksheet['A'+str(row)].value
		while True:		# chat循环
			linenum = worksheet['C'+str(row)].value
			if int(linenum)%2 != 0:
				query = worksheet['D'+str(row)].value
				chat.append(query)
				row += 1
			else:
				query = worksheet['E'+str(row)].value
				candi_qQAs = [query]
				candi_QAs = {}
				while True:		# query循环
					question = worksheet['F'+str(row)].value
					if type(question) == 'long':
						print "Long candi act in row", row
					score = worksheet['I'+str(row)].value
					candi_QAs[question] = [score]
					while True:		# quest循环
						answer = worksheet['J'+str(row)].value
						candi_QAs[question].append(answer)
						row += 1
						# quest循环跳出条件
						question_ = worksheet['F'+str(row)].value
						if question_ != question:
							break
					# query循环跳出条件
					query_ = worksheet['E'+str(row)].value
					if query_ != query:
						break
				candi_qQAs.append(candi_QAs)
				chat.append(candi_qQAs)
			# 退出chat循环的条件
			chatid_ = worksheet['A'+str(row)].value
			if chatid_ != chatid:
				break
		chats.append(chat)
	return chats


# 统计对话数
def cal_chat_num(data_file, sheet_name):
	lines_num = 6899
	workbook = load_workbook(data_file)
	worksheet = workbook[sheet_name]
	chatids = []
	row = 2
	chatid = worksheet['A'+str(row)].value
	chatids.append(chatid)
	while row < lines_num:
		cur_chatid = worksheet['A'+str(row)].value
		if cur_chatid != chatid:
			chatid = cur_chatid
			chatids.append(chatid)
		row += 1
	return chatids

def display_chat(chats):
    for chat in chats:
        i = 0
        for querys in chat:
            if i % 2 != 0:
                print "query:", querys
            else:
                query = querys[0]
                print "query", query
                quests = querys[1]
                for quest in quests.keys():
                    print "    quest", quest
                    answrs = quests[quest]
                    for answr in answrs:
                        print "        answr", answr
            i += 1
        print

def display_context(context):
    state = context[0]
    print "State:",
    for i in range(len(state)):
        print state[i],
    print
    action = context[1]
    print "Action:", action
    reward = context[2]
    print "Reward:", reward
    next_state = context[3]
    print "Next state:",
    for i in range(len(next_state)):
        print next_state[i],
    print


# 去除有缺失的对话段
def del_lack(data_file, sheet_name):
    lines_num = 242293
    rgl_wb = load_workbook(data_file)
    rgl_ws = rgl_wb[sheet_name]
    fnl_wb = Workbook()
    fnl_ws = fnl_wb.active
    fnl_ws.title = sheet_name
    r_row = 1
    f_row = 1
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    for col in cols:
        fnl_ws[col+str(f_row)].value = rgl_ws[col+str(r_row)].value
    r_row += 1
    f_row += 1
    cnt = 0
    while r_row < lines_num:
        # 将数据复制到目标表格
        for col in cols:
            fnl_ws[col+str(f_row)].value = rgl_ws[col+str(r_row)].value
        f_row += 1
        # 判断是否有偶数句缺失
        linum = rgl_ws['C'+str(r_row)].value
        linum_ = rgl_ws['C'+str(r_row+1)].value
        if linum != linum_ and linum%2 == linum_%2:
            cnt += 1
            print r_row, cnt, linum, linum_
            chatid = rgl_ws['A'+str(r_row)].value
            while chatid == rgl_ws['A'+str(r_row)].value:
                r_row += 1
            print r_row
        else:
            r_row += 1
    for col in cols:
        fnl_ws[col+str(f_row)].value = rgl_ws[col+str(r_row)].value
    fnl_wb.save('multi_turn_chat_data_del.xlsx')
            


"""
data_file = 'multi_turn_chat_data_part.xlsx'
sheet_name = 'multi_turn_chats'
chats = fetch_chats(data_file, sheet_name)
#chatids = cal_chat_num(data_file, sheet_name)
print len(chats)
display_chat(chats)
"""
#del_lack('multi_turn_chat_data.xlsx', 'multi_turn_chats')
